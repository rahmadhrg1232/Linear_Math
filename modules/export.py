"""
modules/export.py
==================
Fungsi-fungsi untuk mengekspor hasil perhitungan ke file PDF dan Excel.
Dipisah dari kode UI supaya gampang diuji & dipakai ulang baik dari
halaman Grafik maupun Simpleks.
"""

import io
import os
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ======================================================================
# EKSPOR PDF
# ======================================================================

def export_pdf_grafik(path_file, data_soal, titik_sudut, optimal, gambar_path=None):
    """
    Ekspor hasil Metode Grafik ke PDF: soal, tabel titik sudut, grafik, & solusi.

    data_soal : dict {'obj_type', 'c1', 'c2', 'kendala': [...]}
    titik_sudut : list of (x, y, z)
    optimal : (x, y, z)
    gambar_path : path PNG grafik matplotlib (opsional, akan disisipkan ke PDF)
    """
    doc = SimpleDocTemplate(path_file, pagesize=A4,
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    judul_style = ParagraphStyle('Judul', parent=styles['Title'], fontSize=18)
    elemen = []

    elemen.append(Paragraph("Laporan Program Linear - Metode Grafik", judul_style))
    elemen.append(Paragraph(f"Dibuat: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles['Normal']))
    elemen.append(Spacer(1, 14))

    jenis = "Maksimumkan" if data_soal['obj_type'] == 'max' else "Minimumkan"
    elemen.append(Paragraph(
        f"<b>Fungsi Tujuan:</b> {jenis} Z = {data_soal['c1']}x1 + {data_soal['c2']}x2",
        styles['Normal']))
    elemen.append(Spacer(1, 6))

    elemen.append(Paragraph("<b>Fungsi Batasan:</b>", styles['Normal']))
    for i, k in enumerate(data_soal['kendala'], start=1):
        elemen.append(Paragraph(
            f"K{i}: {k['a1']}x1 + {k['a2']}x2 {k['operator']} {k['rhs']}", styles['Normal']))
    elemen.append(Spacer(1, 14))

    if gambar_path and os.path.exists(gambar_path):
        elemen.append(Paragraph("<b>Grafik Daerah Layak:</b>", styles['Normal']))
        elemen.append(Spacer(1, 6))
        elemen.append(RLImage(gambar_path, width=14 * cm, height=10 * cm))
        elemen.append(Spacer(1, 14))

    elemen.append(Paragraph("<b>Titik-Titik Sudut:</b>", styles['Normal']))
    data_tabel = [["X", "Y", "Z"]] + [[f"{x:.4f}", f"{y:.4f}", f"{z:.4f}"] for x, y, z in titik_sudut]
    tabel = Table(data_tabel, hAlign='LEFT')
    tabel.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F6EF7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elemen.append(tabel)
    elemen.append(Spacer(1, 14))

    if optimal:
        elemen.append(Paragraph(
            f"<b>Solusi Optimal:</b> X = {optimal[0]:.4f}, Y = {optimal[1]:.4f}, "
            f"Z = {optimal[2]:.4f}", styles['Normal']))

    doc.build(elemen)
    return path_file


def export_pdf_simpleks(path_file, data_soal, hasil):
    """
    Ekspor hasil Metode Simpleks ke PDF: soal, tabel tiap iterasi (dengan
    highlight pivot), penjelasan, dan hasil akhir.
    """
    doc = SimpleDocTemplate(path_file, pagesize=A4,
                             topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    judul_style = ParagraphStyle('Judul', parent=styles['Title'], fontSize=18)
    kecil = ParagraphStyle('Kecil', parent=styles['Normal'], fontSize=9, leading=12)
    elemen = []

    elemen.append(Paragraph("Laporan Program Linear - Metode Simpleks", judul_style))
    elemen.append(Paragraph(f"Dibuat: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles['Normal']))
    elemen.append(Spacer(1, 14))

    jenis = "Maksimumkan" if data_soal['obj_type'] == 'max' else "Minimumkan"
    fungsi_z = " + ".join(f"{c}x{i+1}" for i, c in enumerate(data_soal['c']))
    elemen.append(Paragraph(f"<b>Fungsi Tujuan:</b> {jenis} Z = {fungsi_z}", styles['Normal']))
    elemen.append(Spacer(1, 6))
    elemen.append(Paragraph("<b>Fungsi Batasan:</b>", styles['Normal']))
    for i, k in enumerate(data_soal['kendala'], start=1):
        baris = " + ".join(f"{v}x{j+1}" for j, v in enumerate(k['coeffs']))
        elemen.append(Paragraph(f"K{i}: {baris} {k['operator']} {k['rhs']}", styles['Normal']))
    elemen.append(Spacer(1, 16))

    for it in hasil['iterasi']:
        elemen.append(Paragraph(f"<b>Iterasi {it['nomor']}</b>", styles['Heading3']))
        header = ["Basis"] + it['nama_kolom'] + ["NK"]
        data_tabel = [header]
        for r, baris in enumerate(it['tabel']):
            nama_basis = it['nama_kolom'][it['basis_idx'][r]]
            data_tabel.append([nama_basis] + [f"{v:.3f}" for v in baris])
        data_tabel.append(["Cj-Zj"] + [f"{v:.3f}" for v in it['CjZj']] + [""])

        tabel = Table(data_tabel, hAlign='LEFT', repeatRows=1)
        estilo = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F6EF7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]
        if it.get('pivot_kolom') is not None:
            col = it['pivot_kolom'] + 1  # +1 karena kolom 0 = nama basis
            estilo.append(('BACKGROUND', (col, 1), (col, len(it['tabel'])), colors.HexColor('#BFDBFE')))
        if it.get('pivot_baris') is not None:
            row = it['pivot_baris'] + 1  # +1 karena baris 0 = header
            estilo.append(('BACKGROUND', (0, row), (-2, row), colors.HexColor('#FEF08A')))
        if it.get('pivot_kolom') is not None and it.get('pivot_baris') is not None:
            estilo.append(('BACKGROUND', (col, row), (col, row), colors.HexColor('#FCA5A5')))
        tabel.setStyle(TableStyle(estilo))
        elemen.append(tabel)
        elemen.append(Spacer(1, 6))
        elemen.append(Paragraph(it['penjelasan'], kecil))
        elemen.append(Spacer(1, 14))

    elemen.append(Paragraph("<b>Hasil Akhir:</b>", styles['Heading3']))
    if hasil['status'] == 'optimal':
        for nama, val in hasil['nilai_variabel'].items():
            elemen.append(Paragraph(f"{nama} = {val:.4f}", styles['Normal']))
        elemen.append(Paragraph(f"Z optimal = {hasil['Z_optimal']:.4f}", styles['Normal']))
    else:
        elemen.append(Paragraph(f"Status: {hasil['status']}", styles['Normal']))

    doc.build(elemen)
    return path_file


# ======================================================================
# EKSPOR EXCEL
# ======================================================================

_FILL_HEADER = PatternFill("solid", fgColor="4F6EF7")
_FILL_PIVOT_KOLOM = PatternFill("solid", fgColor="BFDBFE")
_FILL_PIVOT_BARIS = PatternFill("solid", fgColor="FEF08A")
_FILL_PIVOT_ELEMEN = PatternFill("solid", fgColor="FCA5A5")
_FONT_HEADER = Font(bold=True, color="FFFFFF")
_BORDER_TIPIS = Border(*(Side(style="thin", color="CCCCCC"),) * 4)


def export_excel_grafik(path_file, data_soal, titik_sudut, optimal):
    """Ekspor hasil Metode Grafik ke Excel (1 sheet: soal + titik sudut + solusi)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Metode Grafik"

    jenis = "Maksimumkan" if data_soal['obj_type'] == 'max' else "Minimumkan"
    ws["A1"] = "Laporan Program Linear - Metode Grafik"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Fungsi Tujuan:"
    ws["B3"] = f"{jenis} Z = {data_soal['c1']}x1 + {data_soal['c2']}x2"

    baris = 5
    ws.cell(row=baris, column=1, value="Fungsi Batasan:").font = Font(bold=True)
    baris += 1
    for i, k in enumerate(data_soal['kendala'], start=1):
        ws.cell(row=baris, column=1, value=f"K{i}: {k['a1']}x1 + {k['a2']}x2 {k['operator']} {k['rhs']}")
        baris += 1

    baris += 1
    header_row = baris
    for col, teks in enumerate(["X", "Y", "Z"], start=1):
        c = ws.cell(row=header_row, column=col, value=teks)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center")
    baris += 1
    for x, y, z in titik_sudut:
        ws.cell(row=baris, column=1, value=round(x, 4))
        ws.cell(row=baris, column=2, value=round(y, 4))
        ws.cell(row=baris, column=3, value=round(z, 4))
        baris += 1

    if optimal:
        baris += 1
        ws.cell(row=baris, column=1, value="Solusi Optimal:").font = Font(bold=True)
        baris += 1
        ws.cell(row=baris, column=1, value=f"X = {optimal[0]:.4f}, Y = {optimal[1]:.4f}, Z = {optimal[2]:.4f}")

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(path_file)
    return path_file


def export_excel_simpleks(path_file, data_soal, hasil):
    """Ekspor hasil Metode Simpleks ke Excel (1 sheet per iterasi + sheet ringkasan)."""
    wb = Workbook()
    ws_ringkasan = wb.active
    ws_ringkasan.title = "Ringkasan"

    jenis = "Maksimumkan" if data_soal['obj_type'] == 'max' else "Minimumkan"
    fungsi_z = " + ".join(f"{c}x{i+1}" for i, c in enumerate(data_soal['c']))
    ws_ringkasan["A1"] = "Laporan Program Linear - Metode Simpleks"
    ws_ringkasan["A1"].font = Font(bold=True, size=14)
    ws_ringkasan["A3"] = "Fungsi Tujuan:"
    ws_ringkasan["B3"] = f"{jenis} Z = {fungsi_z}"

    baris = 5
    for i, k in enumerate(data_soal['kendala'], start=1):
        koef_teks = " + ".join(f"{v}x{j+1}" for j, v in enumerate(k['coeffs']))
        ws_ringkasan.cell(row=baris, column=1, value=f"K{i}: {koef_teks} {k['operator']} {k['rhs']}")
        baris += 1

    baris += 2
    ws_ringkasan.cell(row=baris, column=1, value="Status:").font = Font(bold=True)
    ws_ringkasan.cell(row=baris, column=2, value=hasil['status'])
    baris += 1
    if hasil['status'] == 'optimal':
        for nama, val in hasil['nilai_variabel'].items():
            ws_ringkasan.cell(row=baris, column=1, value=nama)
            ws_ringkasan.cell(row=baris, column=2, value=val)
            baris += 1
        ws_ringkasan.cell(row=baris, column=1, value="Z optimal").font = Font(bold=True)
        ws_ringkasan.cell(row=baris, column=2, value=hasil['Z_optimal'])

    for col in range(1, 4):
        ws_ringkasan.column_dimensions[get_column_letter(col)].width = 24

    # Satu sheet per iterasi, lengkap dengan highlight pivot
    for it in hasil['iterasi']:
        ws = wb.create_sheet(title=f"Iterasi {it['nomor']}")
        header = ["Basis"] + it['nama_kolom'] + ["NK"]
        for col, teks in enumerate(header, start=1):
            c = ws.cell(row=1, column=col, value=teks)
            c.font = _FONT_HEADER
            c.fill = _FILL_HEADER
            c.border = _BORDER_TIPIS
            c.alignment = Alignment(horizontal="center")

        for r, baris_data in enumerate(it['tabel']):
            nama_basis = it['nama_kolom'][it['basis_idx'][r]]
            ws.cell(row=r + 2, column=1, value=nama_basis).border = _BORDER_TIPIS
            for col, val in enumerate(baris_data, start=2):
                cell = ws.cell(row=r + 2, column=col, value=round(float(val), 4))
                cell.border = _BORDER_TIPIS
                if it.get('pivot_kolom') is not None and col - 2 == it['pivot_kolom']:
                    cell.fill = _FILL_PIVOT_KOLOM
                if it.get('pivot_baris') is not None and r == it['pivot_baris']:
                    cell.fill = _FILL_PIVOT_BARIS
                if (it.get('pivot_kolom') is not None and it.get('pivot_baris') is not None
                        and col - 2 == it['pivot_kolom'] and r == it['pivot_baris']):
                    cell.fill = _FILL_PIVOT_ELEMEN

        baris_cjzj = len(it['tabel']) + 2
        ws.cell(row=baris_cjzj, column=1, value="Cj-Zj").font = Font(bold=True)
        for col, val in enumerate(it['CjZj'], start=2):
            ws.cell(row=baris_cjzj, column=col, value=round(float(val), 4))

        ws.cell(row=baris_cjzj + 2, column=1, value="Penjelasan:").font = Font(bold=True, italic=True)
        ws.cell(row=baris_cjzj + 3, column=1, value=it['penjelasan'])
        ws.merge_cells(start_row=baris_cjzj + 3, start_column=1,
                        end_row=baris_cjzj + 3, end_column=max(4, len(header)))
        ws.cell(row=baris_cjzj + 3, column=1).alignment = Alignment(wrap_text=True)

        for col in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    wb.save(path_file)
    return path_file
